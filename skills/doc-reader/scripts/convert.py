#!/usr/bin/env python3
"""
Document Converter — Convert between document formats.

Supported conversions:
  PDF → Text, Markdown, HTML, Images (per-page PNG)
  DOCX → Text, Markdown, PDF, HTML
  XLSX → CSV, JSON, Markdown
  HTML → Text, Markdown, PDF
  Images → Text (OCR), PDF
  Markdown → PDF, HTML, DOCX

Uses: pandoc, pdftotext, PyMuPDF, python-docx, LibreOffice

Usage:
  python3 convert.py <input> <output> [--pages 1-5] [--ocr-lang eng] [--dpi 300]
  python3 convert.py report.pdf report.md
  python3 convert.py data.xlsx data.csv --sheet "Revenue"
  python3 convert.py scan.png scan.txt --ocr-lang eng+nor
"""

import argparse
import csv
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from extract import extract_document, PDFExtractor, _parse_page_range


def convert(input_path: str, output_path: str, pages: str = None,
            ocr_lang: str = "eng", dpi: int = 300, sheet: str = None) -> bool:
    """Convert document from input format to output format."""
    inp = Path(input_path).resolve()
    out = Path(output_path).resolve()
    in_ext = inp.suffix.lower()
    out_ext = out.suffix.lower()

    out.parent.mkdir(parents=True, exist_ok=True)

    # ---- PDF conversions ----
    if in_ext == '.pdf':
        if out_ext in ('.txt', '.md', '.markdown'):
            return _pdf_to_text(inp, out, pages, ocr_lang)
        elif out_ext == '.html':
            return _pdf_to_html(inp, out)
        elif out_ext == '.png':
            return _pdf_to_images(inp, out, pages, dpi)
        elif out_ext == '.json':
            return _doc_to_json(inp, out, pages)

    # ---- DOCX conversions ----
    elif in_ext in ('.docx', '.doc'):
        if out_ext in ('.txt', '.md', '.markdown'):
            return _docx_to_text(inp, out)
        elif out_ext == '.pdf':
            return _pandoc_convert(inp, out)
        elif out_ext == '.html':
            return _pandoc_convert(inp, out)

    # ---- XLSX conversions ----
    elif in_ext in ('.xlsx', '.xls'):
        if out_ext == '.csv':
            return _xlsx_to_csv(inp, out, sheet)
        elif out_ext == '.json':
            return _xlsx_to_json(inp, out, sheet)
        elif out_ext in ('.md', '.txt'):
            return _xlsx_to_markdown(inp, out, sheet)

    # ---- Image conversions ----
    elif in_ext in ('.png', '.jpg', '.jpeg', '.tiff', '.tif', '.bmp'):
        if out_ext == '.txt':
            return _image_to_text(inp, out, ocr_lang)
        elif out_ext == '.pdf':
            return _image_to_pdf(inp, out)

    # ---- HTML conversions ----
    elif in_ext in ('.html', '.htm'):
        if out_ext in ('.txt', '.md'):
            return _html_to_text(inp, out)
        elif out_ext == '.pdf':
            return _pandoc_convert(inp, out)

    # ---- Markdown conversions ----
    elif in_ext in ('.md', '.markdown'):
        if out_ext == '.pdf':
            return _pandoc_convert(inp, out)
        elif out_ext == '.html':
            return _pandoc_convert(inp, out)
        elif out_ext == '.docx':
            return _pandoc_convert(inp, out)

    # ---- CSV conversions ----
    elif in_ext == '.csv':
        if out_ext == '.json':
            return _csv_to_json(inp, out)
        elif out_ext in ('.md', '.txt'):
            return _csv_to_markdown(inp, out)
        elif out_ext == '.xlsx':
            return _csv_to_xlsx(inp, out)

    print(f"Unsupported conversion: {in_ext} → {out_ext}")
    return False


def _pdf_to_text(inp, out, pages, ocr_lang):
    result = extract_document(str(inp), mode="text", pages=pages, ocr_lang=ocr_lang)
    text = result.get("text", "")
    out.write_text(text, encoding='utf-8')
    print(f"✅ {inp.name} → {out.name} ({result.get('stats', {}).get('words', 0)} words)")
    return True


def _pdf_to_html(inp, out):
    try:
        subprocess.run(["pdftohtml", "-s", "-noframes", str(inp), str(out.with_suffix(""))],
                       capture_output=True, timeout=60)
        print(f"✅ {inp.name} → {out.name}")
        return True
    except FileNotFoundError:
        return _pandoc_convert(inp, out)


def _pdf_to_images(inp, out, pages, dpi):
    """Render PDF pages as PNG images."""
    import pymupdf
    doc = pymupdf.open(str(inp))
    page_range = _parse_page_range(pages, doc.page_count) if pages else range(doc.page_count)

    out_dir = out.parent
    stem = out.stem

    saved = []
    for i in page_range:
        page = doc[i]
        pix = page.get_pixmap(dpi=dpi)
        img_path = out_dir / f"{stem}_page{i + 1:03d}.png"
        pix.save(str(img_path))
        saved.append(str(img_path))

    doc.close()
    print(f"✅ {inp.name} → {len(saved)} PNG images in {out_dir}")
    return True


def _doc_to_json(inp, out, pages):
    result = extract_document(str(inp), mode="summary", pages=pages, table_format="json")
    out.write_text(json.dumps(result, indent=2, default=str), encoding='utf-8')
    print(f"✅ {inp.name} → {out.name}")
    return True


def _docx_to_text(inp, out):
    result = extract_document(str(inp), mode="text")
    out.write_text(result.get("text", ""), encoding='utf-8')
    print(f"✅ {inp.name} → {out.name}")
    return True


def _pandoc_convert(inp, out):
    try:
        subprocess.run(["pandoc", str(inp), "-o", str(out)], capture_output=True, timeout=60, check=True)
        print(f"✅ {inp.name} → {out.name} (via pandoc)")
        return True
    except Exception as e:
        print(f"❌ pandoc failed: {e}")
        return False


def _xlsx_to_csv(inp, out, sheet):
    from openpyxl import load_workbook
    wb = load_workbook(str(inp), read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    with open(str(out), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    print(f"✅ {inp.name} → {out.name} (sheet: {ws.title})")
    return True


def _xlsx_to_json(inp, out, sheet):
    from openpyxl import load_workbook
    wb = load_workbook(str(inp), read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        out.write_text("[]", encoding='utf-8')
        return True

    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            record[key] = val
        records.append(record)

    out.write_text(json.dumps(records, indent=2, default=str), encoding='utf-8')
    print(f"✅ {inp.name} → {out.name} ({len(records)} records)")
    return True


def _xlsx_to_markdown(inp, out, sheet):
    result = extract_document(str(inp), mode="text", sheet=sheet)
    out.write_text(result.get("text", ""), encoding='utf-8')
    print(f"✅ {inp.name} → {out.name}")
    return True


def _image_to_text(inp, out, ocr_lang):
    result = extract_document(str(inp), mode="text", ocr_lang=ocr_lang)
    out.write_text(result.get("text", ""), encoding='utf-8')
    print(f"✅ {inp.name} → {out.name} (OCR, lang={ocr_lang})")
    return True


def _image_to_pdf(inp, out):
    import pymupdf
    doc = pymupdf.open()
    img = pymupdf.open(str(inp))
    rect = img[0].rect
    pdf_page = doc.new_page(width=rect.width, height=rect.height)
    pdf_page.insert_image(rect, filename=str(inp))
    doc.save(str(out))
    doc.close()
    print(f"✅ {inp.name} → {out.name}")
    return True


def _html_to_text(inp, out):
    result = extract_document(str(inp), mode="text")
    out.write_text(result.get("text", ""), encoding='utf-8')
    print(f"✅ {inp.name} → {out.name}")
    return True


def _csv_to_json(inp, out):
    with open(str(inp), 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        records = list(reader)
    out.write_text(json.dumps(records, indent=2), encoding='utf-8')
    print(f"✅ {inp.name} → {out.name} ({len(records)} records)")
    return True


def _csv_to_markdown(inp, out):
    result = extract_document(str(inp), mode="text")
    out.write_text(result.get("text", ""), encoding='utf-8')
    print(f"✅ {inp.name} → {out.name}")
    return True


def _csv_to_xlsx(inp, out):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    with open(str(inp), 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
    wb.save(str(out))
    print(f"✅ {inp.name} → {out.name}")
    return True


def main():
    parser = argparse.ArgumentParser(description="Document Converter")
    parser.add_argument('input', help='Input file')
    parser.add_argument('output', help='Output file (format determined by extension)')
    parser.add_argument('--pages', type=str, help='Page range (PDFs): 1-5,8')
    parser.add_argument('--ocr-lang', default='eng', help='OCR language (default: eng)')
    parser.add_argument('--dpi', type=int, default=300, help='DPI for image rendering')
    parser.add_argument('--sheet', type=str, help='Excel sheet name')
    args = parser.parse_args()

    success = convert(args.input, args.output, args.pages, args.ocr_lang, args.dpi, args.sheet)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
